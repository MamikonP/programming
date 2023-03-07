#!/usr/bin/env python3

from openpyxl import load_workbook
import json
import argparse
import sys


def delete_unneccesary_numbers(sheet, maxrow):
    rows = 1
    for i in range(1, maxrow + 1):
        if i == rows:
            rows = rows + 11
            continue
        sheet.cell(row = i, column = 1).value = None

def get_questions_from_sheet(sheet):
    dict_questions = {}
    dict_answers = {}
    questions = []
    question_number = ''
    answer_number = ''
    maxrow = sheet.max_row
    if maxrow == 1:
        print("The current sheet is empty")
        sys.exit(0)
    delete_unneccesary_numbers(sheet, maxrow)
    for i in range(1, maxrow + 1):
        if not sheet.cell(row=i, column=1).value is None:
            if len(questions) != 0:
                dict_questions[question_number] = questions
                questions = []
            number = sheet.cell(row=i, column=1).value - 1
            question_number = 'question_%s' %str(number)
            answer_number = 'answer_%s' %str(number)
            dict_answers[answer_number] = sheet.cell(row=i, column=2).value
        else:
            questions.append(sheet.cell(row = i, column = 2).value)
    dict_questions[question_number] = questions
    return dict_questions, dict_answers

def write_to_json(filename, data):
    with open(filename, 'w') as outfile:
        json.dump(data, outfile, ensure_ascii=False, indent = 4)
    print('created {} file'.format(filename))

def check_sheet_names(sheet_name):
    if sheet_name == 'arm' or sheet_name == 'rus' or sheet_name == 'eng':
        return sheet_name
    print('Default sheet name is \'eng\'')
    return 'eng'


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('Path', metavar='path', type=str,
                        help='the path to excel file')
    parser.add_argument('-s', '--sheet', type=str,
                        help='the sheet name')
    argument = parser.parse_args()
    wb = load_workbook(argument.Path)
    sheet_name = check_sheet_names(argument.sheet)
    sheet = wb[sheet_name]
    dict_questions, dict_answers = get_questions_from_sheet(sheet)
    write_to_json('questions.json', dict_questions)
    write_to_json('answers.json', dict_answers)


if __name__ == '__main__':
    main()
